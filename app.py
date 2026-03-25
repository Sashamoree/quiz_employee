from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory, jsonify
from models import db, Question, UserTest
from database import init_db, get_random_questions, save_test_result, get_statistics
import json
import pandas as pd
from datetime import datetime
import io
from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from datetime import datetime, timedelta
from data import RANKS, EXAMINERS

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///test_sotr.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'pool_recycle': 3600,
    'pool_pre_ping': True,
}

db.init_app(app)

with app.app_context():
    try:
        # Создаем все таблицы
        db.create_all()
        print(" Таблицы созданы")

        # Добавляем тестовые вопросы, если их нет
        if Question.query.count() == 0:
            sample_questions = [
                {
                    'text': 'Что означает аббревиатура "ИБ"?',
                    'opt1': 'Информационная безопасность',
                    'opt2': 'Интернет безопасность',
                    'opt3': 'Информационная база',
                    'correct': 1
                },
                {
                    'text': 'Что такое фишинг?',
                    'opt1': 'Вид компьютерного вируса',
                    'opt2': 'Метод социальной инженерии для получения конфиденциальных данных',
                    'opt3': 'Способ шифрования данных',
                    'correct': 2
                },
                {
                    'text': 'Что из перечисленного является самым надежным паролем?',
                    'opt1': '123456',
                    'opt2': 'password',
                    'opt3': 'qW7#pL2@mR9',
                    'correct': 3
                },
                {
                    'text': 'Что такое двухфакторная аутентификация?',
                    'opt1': 'Вход по двум разным паролям',
                    'opt2': 'Метод защиты, требующий два способа подтверждения личности',
                    'opt3': 'Два администратора для входа',
                    'correct': 2
                },
                {
                    'text': 'Что такое VPN?',
                    'opt1': 'Виртуальная частная сеть для защищенного соединения',
                    'opt2': 'Антивирусная программа',
                    'opt3': 'Тип компьютерного вируса',
                    'correct': 1
                },
                {
                    'text': 'Какое действие не рекомендуется делать с конфиденциальными данными?',
                    'opt1': 'Хранить в зашифрованном виде',
                    'opt2': 'Отправлять по незащищенной электронной почте',
                    'opt3': 'Удалять после использования',
                    'correct': 2
                },
                {
                    'text': 'Что такое социальная инженерия?',
                    'opt1': 'Создание социальных сетей',
                    'opt2': 'Психологическое манипулирование людьми для получения информации',
                    'opt3': 'Инженерное проектирование',
                    'correct': 2
                },
                {
                    'text': 'Как часто рекомендуется менять пароли?',
                    'opt1': 'Раз в 10 лет',
                    'opt2': 'Никогда',
                    'opt3': 'Каждые 30-90 дней',
                    'correct': 3
                },
                {
                    'text': 'Что такое вредоносное ПО (malware)?',
                    'opt1': 'Программа для работы с документами',
                    'opt2': 'Программное обеспечение, наносящее вред компьютеру или данным',
                    'opt3': 'Антивирусная программа',
                    'correct': 2
                },
                {
                    'text': 'Что делать при подозрении на утечку данных?',
                    'opt1': 'Ничего не сообщать',
                    'opt2': 'Немедленно сообщить руководителю и службе ИБ',
                    'opt3': 'Удалить все данные',
                    'correct': 2
                },
                {
                    'text': 'Что такое SSL/TLS сертификат?',
                    'opt1': 'Паспорт пользователя',
                    'opt2': 'Цифровой сертификат для шифрования соединения',
                    'opt3': 'Антивирусная база',
                    'correct': 2
                },
                {
                    'text': 'Как понять, что сайт использует защищенное соединение?',
                    'opt1': 'В адресной строке есть замок и "https://"',
                    'opt2': 'Сайт быстро загружается',
                    'opt3': 'На сайте много рекламы',
                    'correct': 1
                },
                {
                    'text': 'Что такое DDoS-атака?',
                    'opt1': 'Атака с целью кражи паролей',
                    'opt2': 'Перегрузка сервера большим количеством запросов',
                    'opt3': 'Взлом базы данных',
                    'correct': 2
                },
                {
                    'text': 'Что такое "чистый стол"?',
                    'opt1': 'Уборка рабочего места',
                    'opt2': 'Правило, требующее убирать документы с рабочего стола после работы',
                    'opt3': 'Программа для очистки диска',
                    'correct': 2
                },
                {
                    'text': 'Какая информация не является конфиденциальной?',
                    'opt1': 'Пароли сотрудников',
                    'opt2': 'Данные банковских карт',
                    'opt3': 'Публичная информация о компании с официального сайта',
                    'correct': 3
                },
                {
                    'text': 'Что такое "инцидент информационной безопасности"?',
                    'opt1': 'Плановое обновление ПО',
                    'opt2': 'Событие, нарушающее безопасность информации',
                    'opt3': 'Новый сотрудник в отделе',
                    'correct': 2
                },
                {
                    'text': 'Что такое шифрование данных?',
                    'opt1': 'Сжатие данных',
                    'opt2': 'Преобразование данных в нечитаемый вид без ключа',
                    'opt3': 'Копирование данных',
                    'correct': 2
                },
                {
                    'text': 'Что такое бэкап (backup)?',
                    'opt1': 'Удаление данных',
                    'opt2': 'Резервное копирование данных',
                    'opt3': 'Антивирусная проверка',
                    'correct': 2
                },
                {
                    'text': 'Какой пароль считается наиболее безопасным?',
                    'opt1': 'Короткий пароль из цифр',
                    'opt2': 'Длинный пароль из букв, цифр и спецсимволов',
                    'opt3': 'Словарное слово',
                    'correct': 2
                },
                {
                    'text': 'Что такое "безопасный канал связи"?',
                    'opt1': 'Канал с шифрованием данных',
                    'opt2': 'Кабель без помех',
                    'opt3': 'Быстрое интернет-соединение',
                    'correct': 1
                },
                {
                    'text': 'Что делать с найденной на рабочем месте флешкой?',
                    'opt1': 'Вставить в компьютер, чтобы проверить содержимое',
                    'opt2': 'Передать в службу ИБ или руководителю',
                    'opt3': 'Выбросить',
                    'correct': 2
                },
                {
                    'text': 'Что такое "компрометация учетной записи"?',
                    'opt1': 'Создание новой учетной записи',
                    'opt2': 'Несанкционированное получение доступа к учетной записи',
                    'opt3': 'Блокировка учетной записи',
                    'correct': 2
                },
                {
                    'text': 'Какой протокол используется для защищенной передачи данных в интернете?',
                    'opt1': 'HTTP',
                    'opt2': 'HTTPS',
                    'opt3': 'FTP',
                    'correct': 2
                },
                {
                    'text': 'Что такое "защита от социальной инженерии"?',
                    'opt1': 'Антивирусная защита',
                    'opt2': 'Обучение сотрудников распознаванию манипуляций',
                    'opt3': 'Фильтрация трафика',
                    'correct': 2
                },
                {
                    'text': 'Какое правило является основным для работы с конфиденциальными документами?',
                    'opt1': 'Хранить в открытом доступе',
                    'opt2': 'Не передавать третьим лицам без разрешения',
                    'opt3': 'Делиться с коллегами по интересам',
                    'correct': 2
                }
            ]

            for q in sample_questions:
                question = Question(
                    text=q['text'],
                    option1=q['opt1'],
                    option2=q['opt2'],
                    option3=q['opt3'],
                    correct_option=q['correct']
                )
                db.session.add(question)

            db.session.commit()
            print(f" Добавлено {len(sample_questions)} тестовых вопросов")
        else:
            print(f" В базе уже есть {Question.query.count()} вопросов")

    except Exception as e:
        print(f"️ Ошибка при инициализации БД: {e}")
        db.session.rollback()


@app.route('/')
def index():
    return render_template('index.html', ranks=RANKS, examiners=EXAMINERS)

def get_moscow_time():
    return datetime.utcnow()

def format_moscow_time(date_obj):
    if not date_obj:
        return datetime.now().strftime("%H:%M %d.%m.%Y")
    moscow_time = date_obj
    return moscow_time.strftime("%H:%M %d.%m.%Y")


@app.route('/start_test', methods=['POST'])
def start_test():
    try:
        full_name = request.form.get('full_name', '').strip()
        rank = request.form.get('rank', '').strip()
        department = request.form.get('department', '').strip()
        examiner_value = request.form.get('examiner', '').strip()

        print(f"DEBUG: full_name={full_name}")
        print(f"DEBUG: rank={rank}")
        print(f"DEBUG: department={department}")
        print(f"DEBUG: examiner_value={examiner_value}")

        if not full_name:
            flash('Пожалуйста, введите ФИО сотрудника')
            return redirect(url_for('index'))

        if not rank:
            flash('Пожалуйста, выберите звание сотрудника')
            return redirect(url_for('index'))

        if not department:
            flash('Пожалуйста, введите отдел')
            return redirect(url_for('index'))

        if not examiner_value:
            flash('Пожалуйста, выберите экзаменатора')
            return redirect(url_for('index'))

        if '|' in examiner_value:
            examiner_name, examiner_rank = examiner_value.split('|', 1)
        else:
            examiner_name = examiner_value
            examiner_rank = ''

        print(f"DEBUG: examiner_name={examiner_name}, examiner_rank={examiner_rank}")

        questions = get_random_questions(15)

        if len(questions) == 0:
            flash('В базе нет вопросов! Сначала добавьте вопросы в разделе "Вопросы"')
            return redirect(url_for('index'))

        if len(questions) < 15:
            flash(
                f'Внимание! В базе всего {len(questions)} вопросов. Тест будет состоять из {len(questions)} вопросов.')

        return render_template('quiz.html',
                               full_name=full_name,
                               rank=rank,
                               department=department,
                               examiner_name=examiner_name,
                               examiner_rank=examiner_rank,
                               questions=questions)

    except Exception as e:
        print(f"Ошибка в start_test: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Ошибка при начале тестирования: {str(e)}')
        return redirect(url_for('index'))


@app.route('/submit_test', methods=['POST'])
def submit_test():
    try:
        full_name = request.form.get('full_name', '').strip()
        rank = request.form.get('rank', '').strip()
        department = request.form.get('department', '').strip()
        examiner_name = request.form.get('examiner_name', '').strip()
        examiner_rank = request.form.get('examiner_rank', '').strip()

        if not full_name:
            full_name = 'Не указано'
        if not department:
            department = 'Не указано'
        if not examiner_name:
            examiner_name = 'Не указано'

        answers = {}
        correct_count = 0
        questions_data = []

        for key, value in request.form.items():
            if key.startswith('answer_'):
                try:
                    q_id = int(key.split('_')[1])
                    answers[q_id] = int(value)
                except:
                    continue

        all_questions = Question.query.all()
        question_map = {q.id: q for q in all_questions}

        for q_id, answer in answers.items():
            question = question_map.get(q_id)
            if question:
                is_correct = (answer == question.correct_option)
                if is_correct:
                    correct_count += 1
                questions_data.append({
                    'id': q_id,
                    'text': str(question.text),
                    'user_answer': answer,
                    'correct_answer': question.correct_option,
                    'is_correct': is_correct,
                    'options': [
                        str(question.option1),
                        str(question.option2),
                        str(question.option3)
                    ]
                })

        passed = correct_count >= 10

        test_id = save_test_result(full_name, rank, department, examiner_name, examiner_rank, questions_data,
                                   correct_count, passed)

        return render_template('result.html',
                               full_name=full_name,
                               rank=rank,
                               department=department,
                               examiner_name=examiner_name,
                               examiner_rank=examiner_rank,
                               score=correct_count,
                               total=len(questions_data),
                               passed=passed,
                               answers=questions_data,
                               test_id=test_id)

    except Exception as e:
        print(f"Ошибка при обработке теста: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Ошибка при обработке теста: {str(e)}')
        return redirect(url_for('index'))


@app.route('/print_report/<int:test_id>', methods=['GET'])
def print_report(test_id):
    from datetime import datetime, timedelta

    try:
        test = UserTest.query.get_or_404(test_id)

        full_name = test.full_name or 'Не указано'
        rank = test.rank or ''
        department = test.department or 'Не указано'
        examiner_name = test.examiner_name or 'Не указано'
        examiner_rank = test.examiner_rank or ''

        total = test.total_questions if test.total_questions and test.total_questions > 0 else 1
        score = test.score if test.score is not None else 0
        if score > total:
            score = total
        percentage = round(score / total * 100, 1) if total > 0 else 0
        result_text = 'СДАЛ' if test.passed else 'НЕ СДАЛ'

        moscow_now = datetime.utcnow()
        try:
            if test.test_date:
                test_date = test.test_date
                test_date_str = test_date.strftime("%d.%m.%Y %H:%M")
            else:
                test_date_str = moscow_now.strftime("%d.%m.%Y %H:%M")
        except:
            test_date_str = moscow_now.strftime("%d.%m.%Y %H:%M")

        answers = []
        if test.answers_json:
            try:
                answers = json.loads(test.answers_json)
            except:
                answers = []

        max_answers = len(answers)

        html = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Результаты тестирования</title>
    <style>
        @page {
            size: A4;
            margin: 1cm;
        }
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Times New Roman', Times, serif;
            font-size: 10pt;
            line-height: 1.2;
        }
        .container {
            width: 100%;
        }
        h1 {
            text-align: center;
            font-size: 14pt;
            margin-bottom: 8px;
            text-transform: uppercase;
        }
        .info-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 8px;
            border: 1px solid #000;
            font-size: 9pt;
        }
        .info-table td {
            padding: 3px 6px;
            border: 1px solid #000;
            vertical-align: top;
        }
        .info-label {
            font-weight: bold;
            width: 100px;
            background: #f5f5f5;
        }
        .result-block {
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
            margin-bottom: 8px;
        }
        .result-text {
            font-size: 12pt;
            font-weight: bold;
        }
        .result-stats {
            font-size: 9pt;
        }
        .table-wrapper {
            overflow-x: auto;
            margin-bottom: 8px;
        }
        .answers-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 8pt;
            min-width: 600px;
        }
        .answers-table th, .answers-table td {
            border: 1px solid #000;
            padding: 3px 4px;
            vertical-align: top;
        }
        .answers-table th {
            background: #f5f5f5;
            font-weight: bold;
            text-align: center;
        }
        .col-num {
            width: 35px;
            text-align: center;
        }
        .col-question {
            min-width: 250px;
        }
        .col-user {
            min-width: 150px;
        }
        .col-correct {
            min-width: 150px;
        }
        .underline {
            text-decoration: underline;
        }
        .signatures {
            display: flex;
            justify-content: space-between;
            margin-top: 10px;
        }
        .signature-item {
            width: 45%;
        }
        .signature-line {
            border-top: 1px solid #000;
            margin-top: 25px;
            width: 100%;
        }
        .signature-label {
            font-size: 8pt;
            margin-top: 2px;
            text-align: center;
        }
        .date {
            text-align: right;
            margin-top: 8px;
            font-size: 8pt;
        }
        .no-print {
            text-align: center;
            margin-top: 10px;
        }
        button {
            padding: 4px 15px;
            font-size: 10pt;
            cursor: pointer;
        }
        @media print {
            .no-print {
                display: none;
            }
            .table-wrapper {
                overflow-x: visible;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>РЕЗУЛЬТАТЫ ТЕСТИРОВАНИЯ</h1>

        <table class="info-table">
            <tr>
                <td class="info-label">ФИО сотрудника:</td>
                <td colspan="3">''' + full_name[:60] + '''</td>
            </tr>
            <tr>'''

        if rank:
            html += '''
                <td class="info-label">Должность:</td>
                <td>''' + rank[:40] + '''</td>'''
        else:
            html += '<td class="info-label"></td><td></td>'

        html += '''
                <td class="info-label">Отдел:</td>
                <td>''' + department[:40] + '''</td>
            </tr>
            <tr>
                <td class="info-label">Экзаменатор:</td>
                <td>''' + examiner_name[:50] + (f' ({examiner_rank[:30]})' if examiner_rank else '') + '''</td>
                <td class="info-label">Дата:</td>
                <td>''' + test_date_str + '''</td>
            </tr>
        </table>

        <div class="result-block">
            <span class="result-text">РЕЗУЛЬТАТ: ''' + result_text + '''</span>
            <span class="result-stats"> | Правильных ответов: ''' + str(score) + ''' из ''' + str(
            total) + ''' | Процент: ''' + str(percentage) + '''%</span>
        </div>

        <div class="table-wrapper">
            <table class="answers-table">
                <thead>
                    <tr>
                        <th class="col-num">№</th>
                        <th class="col-question">Вопрос</th>
                        <th class="col-user">Ответ сотрудника</th>
                        <th class="col-correct">Правильный ответ</th>
                    </tr>
                </thead>
                <tbody>'''

        for i in range(max_answers):
            ans = answers[i]
            q_text = str(ans.get('text', 'Вопрос'))[:120]
            user_ans = ans.get('user_answer', 1)
            correct_ans = ans.get('correct_answer', 1)
            options = ans.get('options', ['', '', ''])
            is_correct = ans.get('is_correct', False)

            user_text = str(options[user_ans - 1])[:70] if user_ans <= len(options) else ''
            correct_text = str(options[correct_ans - 1])[:70] if correct_ans <= len(options) else ''

            underline_class = 'underline' if not is_correct else ''
            html += '''
                <tr>
                    <td class="col-num" style="text-align:center">''' + str(i + 1) + '''</td>
                    <td class="col-question">''' + q_text + '''</td>
                    <td class="col-user ''' + underline_class + '''">''' + user_text + '''</td>
                    <td class="col-correct">''' + correct_text + '''</td>
                </tr>'''

        html += '''
                </tbody>
            </table>
        </div>

        <div class="signatures">
            <div class="signature-item">
                <div class="signature-line"></div>
                <div class="signature-label">(подпись экзаменатора)</div>
            </div>
            <div class="signature-item">
                <div class="signature-line"></div>
                <div class="signature-label">(подпись сотрудника)</div>
            </div>
        </div>

        <div class="date">
            Дата: ''' + moscow_now.strftime("%d.%m.%Y") + '''
        </div>

        <div class="no-print">
            <button onclick="window.print()">Распечатать отчет</button>
        </div>
    </div>
</body>
</html>'''

        return html

    except Exception as e:
        print(f"Ошибка: {e}")
        flash(f'Ошибка при создании отчета: {str(e)}')
        return redirect(url_for('statistics'))

@app.route('/questions', methods=['GET', 'POST'])
def manage_questions():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Вопросов на странице

    query = Question.query

    if search:
        query = query.filter(Question.text.contains(search))

    pagination = query.order_by(Question.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    questions = pagination.items
    total = pagination.total

    if request.method == 'POST':
        text = request.form.get('text', '').strip()
        option1 = request.form.get('option1', '').strip()
        option2 = request.form.get('option2', '').strip()
        option3 = request.form.get('option3', '').strip()

        if not text or not option1 or not option2 or not option3:
            flash('Все поля должны быть заполнены!', 'error')
            return redirect(url_for('manage_questions'))

        question = Question(
            text=text,
            option1=option1,
            option2=option2,
            option3=option3,
            correct_option=int(request.form.get('correct_option'))
        )
        db.session.add(question)
        db.session.commit()
        flash('Вопрос успешно добавлен!', 'success')
        return redirect(url_for('manage_questions'))

    return render_template('questions.html',
                           questions=questions,
                           pagination=pagination,
                           search=search,
                           total=total,
                           per_page=per_page)


@app.route('/api/question/<int:q_id>')
def get_question_api(q_id):
    try:
        print(f"DEBUG: Запрос вопроса ID: {q_id}")

        question = Question.query.get(q_id)

        if not question:
            print(f"DEBUG: Вопрос с ID {q_id} не найден")
            return jsonify({
                'success': False,
                'error': f'Вопрос с ID {q_id} не найден'
            }), 404

        print(f"DEBUG: Вопрос найден: {question.text[:50]}")

        result = {
            'success': True,
            'question': {
                'id': question.id,
                'text': question.text,
                'option1': question.option1,
                'option2': question.option2,
                'option3': question.option3,
                'correct_option': question.correct_option
            }
        }

        print(f"DEBUG: Отправляем ответ: {result}")
        return jsonify(result)

    except Exception as e:
        print(f"DEBUG: Ошибка в API: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/edit_question', methods=['POST'])
def edit_question():
    try:
        q_id = request.form.get('question_id')
        print(f"DEBUG: Редактирование вопроса ID: {q_id}")

        question = Question.query.get(q_id)
        if not question:
            flash('Вопрос не найден!', 'error')
            return redirect(url_for('manage_questions'))

        # Обновляем данные
        question.text = request.form.get('text', '').strip()
        question.option1 = request.form.get('option1', '').strip()
        question.option2 = request.form.get('option2', '').strip()
        question.option3 = request.form.get('option3', '').strip()
        question.correct_option = int(request.form.get('correct_option'))

        db.session.commit()
        print(f"DEBUG: Вопрос {q_id} обновлен")
        flash('Вопрос успешно обновлен!', 'success')

    except Exception as e:
        print(f"DEBUG: Ошибка при редактировании: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Ошибка при обновлении: {str(e)}', 'error')
        db.session.rollback()

    return redirect(url_for('manage_questions'))


@app.route('/delete_question/<int:q_id>')
def delete_question(q_id):
    question = Question.query.get_or_404(q_id)
    db.session.delete(question)
    db.session.commit()
    flash('Вопрос удален')
    return redirect(url_for('manage_questions'))


@app.route('/statistics', methods=['GET', 'POST'])
def statistics():
    from datetime import timedelta

    full_name = request.args.get('full_name', '') or request.form.get('full_name', '')
    department = request.args.get('department', '') or request.form.get('department', '')
    date_from = request.args.get('date_from', '') or request.form.get('date_from', '')
    date_to = request.args.get('date_to', '') or request.form.get('date_to', '')

    query = UserTest.query

    if full_name:
        query = query.filter(UserTest.full_name.contains(full_name))
    if department:
        query = query.filter(UserTest.department.contains(department))
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(UserTest.test_date >= date_from_obj)
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            query = query.filter(UserTest.test_date <= date_to_obj)
        except:
            pass

    results = query.order_by(UserTest.test_date.desc()).all()

    for r in results:
        if r.test_date:
            local_time = r.test_date
            r.formatted_date = local_time.strftime("%d.%m.%Y %H:%M")
        else:
            r.formatted_date = ''

        total = r.total_questions if r.total_questions and r.total_questions > 0 else 1
        r.percentage = round(r.score / total * 100, 1)

    total_tests = len(results)
    passed_tests = len([r for r in results if r.passed])
    failed_tests = total_tests - passed_tests

    if total_tests > 0:
        success_rate = round(passed_tests / total_tests * 100, 1)
    else:
        success_rate = 0

    stats = {
        'total': total_tests,
        'passed': passed_tests,
        'failed': failed_tests,
        'success_rate': success_rate
    }

    return render_template('statistics.html',
                           results=results,
                           stats=stats,
                           filters={
                               'full_name': full_name,
                               'department': department,
                               'date_from': date_from,
                               'date_to': date_to
                           })


@app.route('/export_excel')
def export_excel():
    from datetime import timedelta
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from urllib.parse import unquote

    try:
        full_name = request.args.get('full_name', '')
        department = request.args.get('department', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        full_name = unquote(full_name) if full_name else ''
        department = unquote(department) if department else ''

        query = UserTest.query

        if full_name:
            query = query.filter(UserTest.full_name.contains(full_name))
        if department:
            query = query.filter(UserTest.department.contains(department))
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(UserTest.test_date >= date_from_obj)
            except:
                pass
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_obj = date_to_obj + timedelta(days=1)
                query = query.filter(UserTest.test_date <= date_to_obj)
            except:
                pass

        results = query.order_by(UserTest.test_date.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"

        headers = ['№', 'ФИО', 'Отдел', 'Дата тестирования', 'Правильных ответов', 'Процент', 'Результат']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, r in enumerate(results, 2):
            full_name_val = str(r.full_name)[:255] if r.full_name else ''
            department_val = str(r.department)[:255] if r.department else ''

            if r.test_date:
                try:
                    local_time = r.test_date + timedelta(hours=3)
                    test_date = local_time.strftime("%d.%m.%Y %H:%M")
                except:
                    test_date = ''
            else:
                test_date = ''

            total = r.total_questions if r.total_questions and r.total_questions > 0 else 1
            score = r.score if r.score is not None else 0

            if total > 65535:
                total = 65535
            if score > 65535:
                score = 65535

            percentage = round(score / total * 100, 1) if total > 0 else 0
            if percentage > 100:
                percentage = 100
            if percentage < 0:
                percentage = 0

            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=full_name_val)
            ws.cell(row=row_idx, column=3, value=department_val)
            ws.cell(row=row_idx, column=4, value=test_date)
            ws.cell(row=row_idx, column=5, value=f'{score}/{total}')
            ws.cell(row=row_idx, column=6, value=f'{percentage}%')
            ws.cell(row=row_idx, column=7, value='Зачет' if r.passed else 'Незачет')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length and length < 100:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'statistics_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Ошибка экспорта: {e}")
        import traceback
        traceback.print_exc()

        import csv
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(['№', 'ФИО', 'Отдел', 'Дата тестирования', 'Правильных ответов', 'Процент', 'Результат'])

        for idx, r in enumerate(results, 1):
            total = r.total_questions if r.total_questions and r.total_questions > 0 else 1
            score = r.score if r.score is not None else 0
            percentage = round(score / total * 100, 1) if total > 0 else 0

            test_date = ''
            if r.test_date:
                try:
                    local_time = r.test_date + timedelta(hours=3)
                    test_date = local_time.strftime("%d.%m.%Y %H:%M")
                except:
                    pass

            writer.writerow([
                idx,
                r.full_name,
                r.department,
                test_date,
                f'{score}/{total}',
                f'{percentage}%',
                'Зачет' if r.passed else 'Незачет'
            ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            as_attachment=True,
            download_name=f'statistics_{datetime.now().strftime("%Y%m%d_%H%M")}.csv',
            mimetype='text/csv'
        )


@app.route('/view_errors/<int:test_id>')
def view_errors(test_id):
    test = UserTest.query.get_or_404(test_id)
    answers = json.loads(test.answers_json)

    errors = [a for a in answers if not a['is_correct']]

    return render_template('errors.html',
                           test=test,
                           errors=errors,
                           total_correct=test.score,
                           total_questions=test.total_questions)


@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static/images', 'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/static/images/favicon.ico')
def favicon_direct():
    return send_from_directory('static/images', 'favicon.ico')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)